# 📂 FILE: app/routers/v1/reports.py
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from datetime import UTC, datetime, timedelta
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import select

from app.database import get_db
from app.models.employee import Employee
from app.models.project import Project
from app.models.task import Task
from app.models.vacation import Vacation
from app.models.sprint import Sprint
from app.models.department import Department
from app.core.deps import get_current_user
from app.core.constants import ROLE_ADMIN, ROLE_MANAGER

router = APIRouter(prefix="/reports", tags=["Reports Export"])


def apply_excel_styling(ws, title_text, headers):
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True

    # 1. Title Block
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = title_text
    title_cell.font = Font(name="Segoe UI", size=16, bold=True, color="1F4E78")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40

    # 2. Metadata Block
    ws["A2"].value = "Báo cáo được tạo tự động lúc:"
    ws["B2"].value = datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S UTC")
    ws["A2"].font = Font(name="Segoe UI", size=9, italic=True, color="595959")
    ws["B2"].font = Font(name="Segoe UI", size=9, italic=True, color="595959")
    ws.row_dimensions[2].height = 20

    # Leave a blank row
    ws.row_dimensions[3].height = 15

    # 4. Headers Row
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[4].height = 28

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = Border(
            bottom=Side(style="medium", color="1F4E78"),
            top=Side(style="thin", color="D9D9D9"),
        )

    # Freeze header row
    ws.freeze_panes = "A5"


def auto_fit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        # Don't size based on merged title row
        for cell in col:
            if cell.row == 1:
                continue
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)

        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


@router.get("/export/{report_type}")
def export_report(
    report_type: str,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    # Only Admin and Manager can export reports
    if current_user.role_id not in (ROLE_ADMIN, ROLE_MANAGER):
        raise HTTPException(
            status_code=403, detail="Unauthorized reports export access"
        )

    wb = openpyxl.Workbook()
    ws = wb.active

    thin_border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )
    zebra_fill = PatternFill(
        start_color="F2F6FA", end_color="F2F6FA", fill_type="solid"
    )

    if report_type == "projects":
        ws.title = "Project Performance"
        headers = [
            "Mã Dự án",
            "Tên Dự án",
            "Trạng thái",
            "Tổng số Task",
            "Task Hoàn thành",
            "Tỷ lệ Hoàn thành (%)",
            "Task Quá hạn",
        ]
        apply_excel_styling(ws, "BÁO CÁO TIẾN ĐỘ VÀ HIỆU SUẤT DỰ ÁN", headers)

        projects = db.scalars(select(Project).where(Project.is_deleted == False)).all()
        row_num = 5

        for p in projects:
            tasks = db.scalars(
                select(Task).where(Task.project_id == p.id, Task.is_deleted == False)
            ).all()
            total_tasks = len(tasks)
            completed_tasks = sum(1 for t in tasks if t.status == "Done")
            progress = (
                round((completed_tasks / total_tasks) * 100) if total_tasks > 0 else 0
            )
            current_utc = datetime.now(UTC).replace(tzinfo=None)
            overdue_tasks = sum(
                1
                for t in tasks
                if t.status != "Done" and t.deadline and t.deadline < current_utc
            )

            row_data = [
                p.project_code,
                p.name,
                p.status,
                total_tasks,
                completed_tasks,
                progress,
                overdue_tasks,
            ]

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = Font(name="Segoe UI", size=10)
                cell.border = thin_border
                if row_num % 2 == 0:
                    cell.fill = zebra_fill
                if col_idx in (4, 5, 6, 7):
                    cell.alignment = Alignment(horizontal="right")

            row_num += 1

        # Summary box
        summary_start = row_num + 2
        ws.cell(row=summary_start, column=1, value="Tổng số dự án:").font = Font(
            name="Segoe UI", size=10, bold=True
        )
        ws.cell(row=summary_start, column=2, value=len(projects)).font = Font(
            name="Segoe UI", size=10
        )

        auto_fit_columns(ws)

    elif report_type == "tasks":
        ws.title = "Tasks Distribution"
        headers = [
            "Task ID",
            "Tiêu đề",
            "Dự án",
            "Người thực hiện",
            "Trạng thái",
            "Độ ưu tiên",
            "Story Points",
            "Hạn chót",
        ]
        apply_excel_styling(ws, "BÁO CÁO PHÂN BỔ VÀ TIẾN ĐỘ CÔNG VIỆC", headers)

        tasks = db.scalars(select(Task).where(Task.is_deleted == False)).all()
        row_num = 5

        for t in tasks:
            proj = db.get(Project, t.project_id)
            proj_name = proj.name if proj else "—"

            # Fetch assignee
            assignee_name = "Chưa phân công"
            from app.models.task_assignment import TaskAssignment

            assignee_id = db.scalar(
                select(TaskAssignment.employee_id).where(TaskAssignment.task_id == t.id)
            )
            if assignee_id:
                emp = db.get(Employee, assignee_id)
                if emp:
                    assignee_name = emp.full_name

            deadline_str = t.deadline.strftime("%Y-%m-%d") if t.deadline else "—"
            row_data = [
                t.id,
                t.title,
                proj_name,
                assignee_name,
                t.status,
                t.priority,
                t.story_points,
                deadline_str,
            ]

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = Font(name="Segoe UI", size=10)
                cell.border = thin_border
                if row_num % 2 == 0:
                    cell.fill = zebra_fill
                if col_idx == 7:
                    cell.alignment = Alignment(horizontal="right")

            row_num += 1

        # Add Pie Chart for Status
        if len(tasks) > 0:
            ws.cell(row=row_num + 2, column=1, value="Thống kê Trạng thái:").font = (
                Font(name="Segoe UI", size=11, bold=True)
            )
            todo_c = sum(1 for t in tasks if t.status == "To Do")
            ip_c = sum(1 for t in tasks if t.status == "In Progress")
            done_c = sum(1 for t in tasks if t.status == "Done")

            ws.cell(row=row_num + 3, column=1, value="To Do")
            ws.cell(row=row_num + 3, column=2, value=todo_c)
            ws.cell(row=row_num + 4, column=1, value="In Progress")
            ws.cell(row=row_num + 4, column=2, value=ip_c)
            ws.cell(row=row_num + 5, column=1, value="Done")
            ws.cell(row=row_num + 5, column=2, value=done_c)

            chart = PieChart()
            labels = Reference(ws, min_col=1, min_row=row_num + 3, max_row=row_num + 5)
            data = Reference(ws, min_col=2, min_row=row_num + 3, max_row=row_num + 5)
            chart.add_data(data)
            chart.set_categories(labels)
            chart.title = "Biểu đồ Trạng thái Task"
            ws.add_chart(chart, "D" + str(row_num + 3))

        auto_fit_columns(ws)

    elif report_type == "workload":
        ws.title = "Employee Workload"
        headers = [
            "Nhân viên",
            "Phòng ban",
            "Số Task đang làm",
            "Task Quá hạn",
            "Task Ưu tiên cao",
        ]
        apply_excel_styling(ws, "BÁO CÁO PHÂN BỔ KHỐI LƯỢNG CÔNG VIỆC NHÂN SỰ", headers)

        employees = db.scalars(
            select(Employee).where(Employee.is_deleted == False)
        ).all()
        row_num = 5

        for emp in employees:
            dept = db.get(Department, emp.department_id) if emp.department_id else None
            dept_name = dept.name if dept else "—"

            # Count tasks
            from app.models.task_assignment import TaskAssignment

            emp_task_ids = db.scalars(
                select(TaskAssignment.task_id).where(
                    TaskAssignment.employee_id == emp.id
                )
            ).all()

            active_tasks = 0
            overdue_tasks = 0
            high_priority = 0

            if emp_task_ids:
                tasks = db.scalars(
                    select(Task).where(
                        Task.id.in_(emp_task_ids), Task.is_deleted == False
                    )
                ).all()
                active_tasks = sum(1 for t in tasks if t.status != "Done")
                current_utc = datetime.now(UTC).replace(tzinfo=None)
                overdue_tasks = sum(
                    1
                    for t in tasks
                    if t.status != "Done" and t.deadline and t.deadline < current_utc
                )
                high_priority = sum(
                    1 for t in tasks if t.priority in ("High", "Urgent")
                )

            row_data = [
                emp.full_name,
                dept_name,
                active_tasks,
                overdue_tasks,
                high_priority,
            ]

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = Font(name="Segoe UI", size=10)
                cell.border = thin_border
                if row_num % 2 == 0:
                    cell.fill = zebra_fill
                if col_idx in (3, 4, 5):
                    cell.alignment = Alignment(horizontal="right")

            row_num += 1

        auto_fit_columns(ws)

    elif report_type == "vacations":
        ws.title = "Vacations Summary"
        headers = [
            "Nhân viên",
            "Email",
            "Từ ngày",
            "Đến ngày",
            "Số ngày nghỉ",
            "Loại nghỉ",
            "Trạng thái",
        ]
        apply_excel_styling(ws, "BÁO CÁO TỔNG HỢP NGHỈ PHÉP NHÂN SỰ", headers)

        vacations = db.scalars(
            select(Vacation).where(Vacation.is_deleted == False)
        ).all()
        row_num = 5

        for vac in vacations:
            emp = db.get(Employee, vac.employee_id)
            emp_name = emp.full_name if emp else "—"
            emp_email = emp.email if emp else "—"

            start_str = vac.start_date.strftime("%Y-%m-%d") if vac.start_date else "—"
            end_str = vac.end_date.strftime("%Y-%m-%d") if vac.end_date else "—"
            duration = (
                (vac.end_date - vac.start_date).days + 1
                if (vac.end_date and vac.start_date)
                else 0
            )

            row_data = [
                emp_name,
                emp_email,
                start_str,
                end_str,
                duration,
                vac.type,
                vac.status,
            ]

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = Font(name="Segoe UI", size=10)
                cell.border = thin_border
                if row_num % 2 == 0:
                    cell.fill = zebra_fill
                if col_idx == 5:
                    cell.alignment = Alignment(horizontal="right")

            row_num += 1

        auto_fit_columns(ws)

    elif report_type == "sprints":
        ws.title = "Sprints Overview"
        headers = [
            "Tên Sprint",
            "Mục tiêu",
            "Bắt đầu",
            "Kết thúc",
            "Trạng thái",
            "Capacity",
            "Task Hoàn thành / Tổng",
        ]
        apply_excel_styling(ws, "BÁO CÁO CHU KỲ PHÁT TRIỂN (SPRINTS)", headers)

        sprints = db.scalars(select(Sprint).where(Sprint.is_deleted == False)).all()
        row_num = 5

        for s in sprints:
            start_str = s.start_date.strftime("%Y-%m-%d") if s.start_date else "—"
            end_str = s.end_date.strftime("%Y-%m-%d") if s.end_date else "—"

            tasks = db.scalars(
                select(Task).where(Task.sprint_id == s.id, Task.is_deleted == False)
            ).all()
            total_tasks = len(tasks)
            completed_tasks = sum(1 for t in tasks if t.status == "Done")
            task_ratio = f"{completed_tasks} / {total_tasks}"

            row_data = [
                s.name,
                s.goal,
                start_str,
                end_str,
                s.status,
                s.capacity,
                task_ratio,
            ]

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = Font(name="Segoe UI", size=10)
                cell.border = thin_border
                if row_num % 2 == 0:
                    cell.fill = zebra_fill
                if col_idx == 6:
                    cell.alignment = Alignment(horizontal="right")

            row_num += 1

        auto_fit_columns(ws)

    else:
        raise HTTPException(status_code=400, detail="Invalid report type")

    # Save to memory stream
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    filename = f"{report_type}_report_{datetime.now(UTC).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
