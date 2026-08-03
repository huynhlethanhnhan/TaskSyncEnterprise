# 📂 FILE: tests/test_gap_remediation.py
import pytest
from datetime import datetime, timedelta
from app.models.employee import Employee
from app.models.project import Project
from app.models.task import Task
from app.models.task_assignment import TaskAssignment
from app.models.project_member import ProjectMember
from app.models.sprint import Sprint
from app.models.backlog_item import BacklogItem
from app.models.task_checklist import TaskChecklist
from app.models.task_comment import TaskComment
from app.models.user_feedback import UserFeedback
from app.core.security import get_password_hash
from app.core.constants import ROLE_ADMIN, ROLE_MANAGER, ROLE_EMPLOYEE


def create_user(db, email, role_id):
    hashed = get_password_hash("pass123")
    user = Employee(
        employee_code=f"CODE_{email.split('@')[0]}",
        full_name=f"User {email.split('@')[0]}",
        email=email,
        password_hash=hashed,
        role_id=role_id,
        is_active=True,
        is_deleted=False,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


def get_headers(client, email):
    response = client.post(
        "/api/v1/auth/login", data={"username": email, "password": "pass123"}
    )
    token = response.json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture
def setup_data(db):
    admin = create_user(db, "adm@test.com", ROLE_ADMIN)
    manager = create_user(db, "mgr@test.com", ROLE_MANAGER)
    employee = create_user(db, "emp@test.com", ROLE_EMPLOYEE)
    employee2 = create_user(db, "emp2@test.com", ROLE_EMPLOYEE)

    project = Project(
        name="Gap Project",
        project_code="GAPPRJ",
        status="Active",
        priority="High",
        is_deleted=False,
    )
    db.add(project)
    db.commit()
    db.refresh(project)

    # Scope both the delivery employee and the manager to this project.
    db.add_all(
        [
            ProjectMember(project_id=project.id, employee_id=employee.id),
            ProjectMember(project_id=project.id, employee_id=manager.id),
        ]
    )
    db.commit()

    task = Task(
        project_id=project.id,
        title="Checklist Task",
        status="To Do",
        priority="Medium",
        story_points=3,
        is_deleted=False,
    )
    db.add(task)
    db.commit()
    db.refresh(task)

    # Assign task to employee
    assignment = TaskAssignment(task_id=task.id, employee_id=employee.id)
    db.add(assignment)
    db.commit()

    return {
        "admin": admin,
        "manager": manager,
        "employee": employee,
        "employee2": employee2,
        "project": project,
        "task": task,
    }


def test_checklist_crud(client, db, setup_data):
    # Retrieve auth headers
    emp_headers = get_headers(client, "emp@test.com")
    emp2_headers = get_headers(client, "emp2@test.com")
    task_id = setup_data["task"].id

    # 1. Create checklist item
    response = client.post(
        f"/api/v1/tasks/{task_id}/checklist",
        json={"title": "Write unit tests", "is_completed": False},
        headers=emp_headers,
    )
    assert response.status_code == 201
    item_id = response.json()["id"]

    # IDOR check: employee2 not in project/task should get 403
    response_idor = client.post(
        f"/api/v1/tasks/{task_id}/checklist",
        json={"title": "Malicious check"},
        headers=emp2_headers,
    )
    assert response_idor.status_code == 403

    # 2. Get checklist items
    response = client.get(f"/api/v1/tasks/{task_id}/checklist", headers=emp_headers)
    assert response.status_code == 200
    assert len(response.json()) == 1
    assert response.json()[0]["title"] == "Write unit tests"

    # 3. Update checklist item
    response = client.patch(
        f"/api/v1/tasks/{task_id}/checklist/{item_id}",
        json={"is_completed": True},
        headers=emp_headers,
    )
    assert response.status_code == 200
    assert response.json()["is_completed"] is True

    # 4. Delete checklist item
    response = client.delete(
        f"/api/v1/tasks/{task_id}/checklist/{item_id}", headers=emp_headers
    )
    assert response.status_code == 200


def test_comments_crud(client, db, setup_data):
    emp_headers = get_headers(client, "emp@test.com")
    emp2_headers = get_headers(client, "emp2@test.com")
    task_id = setup_data["task"].id

    # 1. Create comment
    response = client.post(
        f"/api/v1/tasks/{task_id}/comments",
        json={"content": "I am working on this task"},
        headers=emp_headers,
    )
    assert response.status_code == 201
    comment_id = response.json()["id"]

    # 2. Retrieve comments
    response = client.get(f"/api/v1/tasks/{task_id}/comments", headers=emp_headers)
    assert response.status_code == 200
    assert len(response.json()) >= 1
    assert response.json()[0]["content"] == "I am working on this task"

    # 3. Edit comment (author permission check)
    response_err = client.patch(
        f"/api/v1/tasks/{task_id}/comments/{comment_id}",
        json={"content": "Hacked content"},
        headers=emp2_headers,
    )
    assert response_err.status_code == 403

    response = client.patch(
        f"/api/v1/tasks/{task_id}/comments/{comment_id}",
        json={"content": "Edited my comment"},
        headers=emp_headers,
    )
    assert response.status_code == 200

    # 4. Delete comment
    response = client.delete(
        f"/api/v1/tasks/{task_id}/comments/{comment_id}", headers=emp_headers
    )
    assert response.status_code == 200


def test_sprint_lifecycle_and_backlog(client, db, setup_data):
    mgr_headers = get_headers(client, "mgr@test.com")
    emp_headers = get_headers(client, "emp@test.com")
    proj_id = setup_data["project"].id

    # 1. Create backlog item (managers/admins only)
    response = client.post(
        "/api/v1/backlog",
        json={
            "project_id": proj_id,
            "title": "A new feature story",
            "description": "Requires sprint analytics",
            "priority": "High",
            "story_points": 5,
        },
        headers=mgr_headers,
    )
    assert response.status_code == 201
    backlog_id = response.json()["id"]

    # Regular employee try to create backlog item (forbidden)
    response_emp = client.post(
        "/api/v1/backlog",
        json={"project_id": proj_id, "title": "Employee story"},
        headers=emp_headers,
    )
    assert response_emp.status_code == 403

    # 2. Create sprint
    response = client.post(
        "/api/v1/sprints",
        json={
            "project_id": proj_id,
            "name": "Sprint 1",
            "goal": "Deliver core features",
            "capacity": 10,
        },
        headers=mgr_headers,
    )
    assert response.status_code == 201
    sprint_id = response.json()["id"]

    # 3. Associate backlog item with sprint
    response = client.put(
        f"/api/v1/backlog/{backlog_id}",
        json={"sprint_id": sprint_id},
        headers=mgr_headers,
    )
    assert response.status_code == 200
    assert response.json()["sprint_id"] == sprint_id

    # 4. Convert backlog item to task
    response = client.post(
        f"/api/v1/backlog/{backlog_id}/convert-to-task", headers=mgr_headers
    )
    assert response.status_code == 200
    task_id = response.json()["id"]
    assert response.json()["story_points"] == 5

    # 5. Start sprint
    response = client.patch(f"/api/v1/sprints/{sprint_id}/start", headers=mgr_headers)
    assert response.status_code == 200
    assert response.json()["status"] == "Active"

    # 6. Sprint overlap check (cannot start another active sprint for the same project)
    sprint2_response = client.post(
        "/api/v1/sprints",
        json={"project_id": proj_id, "name": "Sprint 2"},
        headers=mgr_headers,
    )
    sprint2_id = sprint2_response.json()["id"]

    response_overlap = client.patch(
        f"/api/v1/sprints/{sprint2_id}/start", headers=mgr_headers
    )
    assert response_overlap.status_code == 409

    # 7. Complete sprint
    response = client.patch(
        f"/api/v1/sprints/{sprint_id}/complete", headers=mgr_headers
    )
    assert response.status_code == 200
    assert response.json()["status"] == "Completed"


def test_backlog_creation_normalizes_empty_optional_fields(client, db, setup_data):
    mgr_headers = get_headers(client, "mgr@test.com")
    proj_id = setup_data["project"].id

    response = client.post(
        "/api/v1/backlog",
        json={
            "project_id": proj_id,
            "title": "A backlog item",
            "description": "",
            "priority": "   ",
            "story_points": "",
            "sprint_id": "",
            "topic_id": "",
        },
        headers=mgr_headers,
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["description"] is None
    assert payload["priority"] == "Medium"
    assert payload["story_points"] == 0
    assert payload["sprint_id"] is None
    assert payload["topic_id"] is None


def test_feedback_module(client, db, setup_data):
    emp_headers = get_headers(client, "emp@test.com")
    mgr_headers = get_headers(client, "mgr@test.com")

    # 1. Submit feedback (employee)
    response = client.post(
        "/api/v1/feedback",
        json={
            "title": "Fix coffee machine",
            "category": "Culture",
            "description": "It has been broken for 3 days.",
            "impact_level": "Medium",
            "is_anonymous": True,
        },
        headers=emp_headers,
    )
    assert response.status_code == 201
    feedback_id = response.json()["id"]
    assert response.json()["is_anonymous"] is True
    # Anonymity check: Employee cannot see the submitter details of an anonymous feedback
    assert response.json()["submitter_id"] is None

    # 2. Get my feedback
    response = client.get("/api/v1/feedback/my", headers=emp_headers)
    assert response.status_code == 200
    assert len(response.json()) == 1

    # 3. Review/Resolve feedback (manager)
    response = client.patch(
        f"/api/v1/feedback/{feedback_id}/review",
        json={"status": "Planned", "response": "We will order parts today."},
        headers=mgr_headers,
    )
    assert response.status_code == 200
    assert response.json()["status"] == "Planned"
    assert response.json()["response"] == "We will order parts today."


def test_excel_export_integrity(client, db, setup_data):
    mgr_headers = get_headers(client, "mgr@test.com")

    # Export projects report
    response = client.get("/api/v1/reports/export/projects", headers=mgr_headers)
    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Export tasks report
    response = client.get("/api/v1/reports/export/tasks", headers=mgr_headers)
    assert response.status_code == 200

    # Load into openpyxl to ensure no corruption
    import io
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    assert "Tasks Distribution" in wb.sheetnames
    ws = wb["Tasks Distribution"]
    assert ws["A1"].value == "BÁO CÁO PHÂN BỔ VÀ TIẾN ĐỘ CÔNG VIỆC"
    assert ws.cell(row=4, column=1).value == "Task ID"


def test_files_aggregation_and_download(client, db, setup_data):
    emp_headers = get_headers(client, "emp@test.com")
    emp2_headers = get_headers(client, "emp2@test.com")

    # Upload a file
    import io

    file_data = {"file": ("test.txt", io.BytesIO(b"Hello world"), "text/plain")}
    response = client.post(
        "/api/v1/files/upload",
        files=file_data,
        data={"task_id": setup_data["task"].id},
        headers=emp_headers,
    )
    assert response.status_code == 200
    file_id = response.json()["data"]["id"]

    # 1. Access File Registry
    response = client.get("/api/v1/files", headers=emp_headers)
    assert response.status_code == 200
    assert any(f["id"] == file_id for f in response.json())

    # IDOR / unauthorized checks
    # Employee 2 (not project member) should not see this file
    response2 = client.get("/api/v1/files", headers=emp2_headers)
    assert not any(f["id"] == file_id for f in response2.json())

    # 2. Download endpoint authorization check (IDOR check)
    dl_response = client.get(f"/api/v1/files/download/{file_id}", headers=emp2_headers)
    assert dl_response.status_code == 403


def test_topics_and_replies(client, db, setup_data):
    emp_headers = get_headers(client, "emp@test.com")
    emp2_headers = get_headers(client, "emp2@test.com")
    proj_id = setup_data["project"].id

    # Create topic
    response = client.post(
        "/api/v1/topics",
        json={
            "project_id": proj_id,
            "title": "Design Discussion",
            "content": "How should we design this?",
        },
        headers=emp_headers,
    )
    assert response.status_code == 201
    topic_id = response.json()["id"]

    # IDOR project membership check
    response_idor = client.post(
        "/api/v1/topics",
        json={
            "project_id": proj_id,
            "title": "Malicious Topic",
            "content": "No access",
        },
        headers=emp2_headers,
    )
    assert response_idor.status_code == 403

    # Add reply
    reply_response = client.post(
        f"/api/v1/topics/{topic_id}/replies",
        json={"content": "I suggest using a schema"},
        headers=emp_headers,
    )
    assert reply_response.status_code == 201
